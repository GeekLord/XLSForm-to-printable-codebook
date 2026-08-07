"""Throwaway generator for a multi-language Kobo-style XLSForm used to verify output."""
import openpyxl

wb = openpyxl.Workbook()

survey = wb.active
survey.title = 'survey'
survey.append([
    'type', 'name', 'label::English (en)', 'label::Hindi (hi)',
    'hint::English (en)', 'hint::Hindi (hi)', 'required', 'relevant',
    'constraint', 'constraint_message::English (en)', 'calculation', 'appearance',
])

rows = [
    ['start', 'starttime', '', '', '', '', '', '', '', '', '', ''],
    ['end', 'endtime', '', '', '', '', '', '', '', '', '', ''],
    ['today', 'today', '', '', '', '', '', '', '', '', '', ''],
    ['deviceid', 'deviceid', '', '', '', '', '', '', '', '', '', ''],
    ['username', 'username', '', '', '', '', '', '', '', '', '', ''],
    ['audit', 'audit', '', '', '', '', '', '', '', '', '', ''],
    ['begin group', 'hh_roster_intro', 'A. Household identification', 'क. परिवार की पहचान', '', '', '', '', '', '', '', 'field-list'],
    ['calculate', 'hh_name_pre', '', '', '', '', '', '', '', '',
     'pulldata(\'household_list\', \'hh_head_name\', \'hh_id\', ${hh_id})', ''],
    ['calculate', 'survey_started', '', '', '', '', '', '', '', '', 'once(now())', ''],
    ['calculate', 'district_label', '', '', '', '', '', '', '', '',
     "jr:choice-name(${district}, '${district}')", ''],
    ['calculate', 'interview_minutes', '', '', '', '', '', '', '', '',
     'if(${endtime} != \'\', (decimal-date-time(${endtime}) - decimal-date-time(${starttime})) * 1440, 0)', ''],
    ['text', 'hh_id', 'Household ID', 'परिवार आईडी', 'Copy from the listing sheet', 'सूची पत्रक से लिखें',
     'yes', '', 'regex(., \'^[0-9]{6}$\')', 'Enter exactly six digits', '', ''],
    ['select_one district', 'district', 'District of residence', 'निवास का जिला', 'Read options aloud',
     'विकल्प पढ़कर सुनाएँ', 'yes', '', '', '', '', 'minimal'],
    ['select_multiple assets', 'assets', 'Which of these assets does the household own?',
     'इनमें से कौन-कौन सी संपत्तियाँ परिवार के पास हैं?', 'Select all that apply', 'सभी लागू विकल्प चुनें',
     '', '${hh_id} != \'\'', '', '', '', ''],
    ['end group', 'hh_roster_intro', '', '', '', '', '', '', '', '', '', ''],
    ['begin repeat', 'child_roster', 'B. Child roster', 'ख. बाल सूची', '', '', '', '', '', '', '', ''],
    ['text', 'child_name', 'Name of child', 'बच्चे का नाम', '', '', 'yes', '', '', '', '', ''],
    ['integer', 'child_age', 'Age of child in completed years', 'बच्चे की पूर्ण आयु (वर्ष)', '', '',
     'yes', '', '. >= 0 and . <= 18', 'Age must be between 0 and 18', '', ''],
    ['select_one yes_no', 'child_enrolled', 'Is the child currently enrolled in school?',
     'क्या बच्चा वर्तमान में स्कूल में नामांकित है?', '', '', 'yes', '${child_age} >= 3', '', '', '', ''],
    ['calculate', 'child_label', '', '', '', '', '', '', '', '',
     'concat(substr(${child_name}, 0, 12), \' - \', string-length(${child_name}))', ''],
    ['end repeat', 'child_roster', '', '', '', '', '', '', '', '', '', ''],
    ['note', 'closing_note', 'Thank the respondent and end the interview.',
     'उत्तरदाता को धन्यवाद दें और साक्षात्कार समाप्त करें।', '', '', '', '', '', '', '', ''],
]
for row in rows:
    survey.append(row)

choices = wb.create_sheet('choices')
choices.append(['list_name', 'name', 'label::English (en)', 'label::Hindi (hi)'])
choice_rows = [
    ['district', '1', 'Khordha', 'खोरधा'],
    ['district', '2', 'Cuttack', 'कटक'],
    ['district', '3', 'Balasore', 'बालासोर'],
    ['assets', '1', 'Bicycle', 'साइकिल'],
    ['assets', '2', 'Mobile phone with internet access', 'इंटरनेट सुविधा वाला मोबाइल फोन'],
    ['assets', '3', 'Refrigerator', 'रेफ्रिजरेटर'],
    ['assets', '96', 'None of the above', 'इनमें से कोई नहीं'],
    ['yes_no', '1', 'Yes', 'हाँ'],
    ['yes_no', '0', 'No', 'नहीं'],
]
for row in choice_rows:
    choices.append(row)

settings = wb.create_sheet('settings')
settings.append(['form_title', 'form_id', 'version', 'default_language'])
settings.append(['Sample Household Survey', 'sample_hh_survey', '2026080801', 'English (en)'])

wb.save('Sample Household Survey.xlsx')
print('wrote Sample Household Survey.xlsx')
