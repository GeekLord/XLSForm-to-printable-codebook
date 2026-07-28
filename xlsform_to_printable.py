import os
import sys
import re
import argparse
import asyncio
from bs4 import BeautifulSoup
import openpyxl
from playwright.async_api import async_playwright

RESPONSIVE_STYLES = """
.spacer {
    display: none !important;
    width: 0 !important;
    padding: 0 !important;
    margin: 0 !important;
}

*, *::before, *::after { box-sizing: border-box; }

body {
    margin: 0;
    padding: 10px;
    font-family: "Arial Unicode MS", "Arial Unicode", Arial, sans-serif;
    font-size: 14px;
    line-height: 20px;
    color: #000;
    background: #fff;
}

.codebook-container {
    width: 100%;
    max-width: 1200px;
    margin: auto;
    padding: 0 15px;
}

h4 {
    text-align: center;
    font-size: 18px;
    margin: 15px 0;
    padding-bottom: 10px;
    border-bottom: 2px solid #333;
}

table {
    width: 100%;
    border-collapse: collapse;
    margin-bottom: 20px;
}

th, td {
    padding: 8px;
    text-align: left;
    border: 1px solid #ddd;
    vertical-align: top;
}

td[colspan="3"],
.section-header-dark {
    background: #1a1a1a !important;
    color: #fff !important;
    font-weight: bold;
    -webkit-print-color-adjust: exact !important;
    print-color-adjust: exact !important;
}

.headerCell {
    background: #707070;
    color: #fff;
}

.headerCell h6 {
    margin: 0;
    font-size: 12px;
    font-weight: bold;
}

.fieldCell {
    font-weight: bold;
    color: #333;
    width: 20%;
    white-space: normal;
    word-break: break-word;
}

.fieldCell .required {
    color: #c00;
    font-weight: normal;
}

.questionCell {
    width: 60%;
}

.answerCell {
    width: 20%;
}

tr.gray td {
    background: #fff;
}

td .table {
    margin: 0;
    font-size: 13px;
}

td .table td {
    padding: 4px;
    border: none;
    border-bottom: 1px dotted #ccc;
}

.response-note-cell {
    width: 10px;
}

@media screen and (max-width: 768px) {
    body {
        font-size: 14px;
        line-height: 22px;
        padding: 5px;
    }
    
    h4 {
        font-size: 16px;
    }
    
    .table-responsive {
        overflow-x: auto;
        -webkit-overflow-scrolling: touch;
    }
    
    table {
        min-width: 600px;
    }
    
    th, td {
        padding: 6px;
    }
    
    .headerCell h6 {
        font-size: 11px;
    }
}

@media screen and (max-width: 480px) {
    body {
        font-size: 13px;
        line-height: 20px;
        padding: 5px;
    }
    
    h4 {
        font-size: 14px;
    }
    
    th, td {
        padding: 4px;
        font-size: 12px;
    }
    
    .fieldCell {
        width: 25%;
    }
    
    .questionCell {
        width: 50%;
    }
    
    .answerCell {
        width: 25%;
    }
}

@media print {
    @page { size: A4 landscape; margin: 10mm; }
    @page :first { margin-top: 8mm; }
    
    body {
        font-size: 9pt;
        line-height: 14pt;
        padding: 0;
        background: #fff;
    }
    
    h4 {
        font-size: 14pt;
        margin: 10px 0;
        border-bottom: 1pt solid #333;
    }
    
    table {
        page-break-inside: auto;
    }
    
    tr {
        page-break-inside: avoid;
    }
    
    th, td {
        padding: 4pt 6pt;
        border: 0.5pt solid #999;
    }
    
    .headerCell {
        background: #1a1a1a !important;
        color: #fff !important;
        -webkit-print-color-adjust: exact;
        print-color-adjust: exact;
    }
    
    td[colspan="3"],
    .section-header-dark {
        background: #1a1a1a !important;
        color: #fff !important;
        -webkit-print-color-adjust: exact !important;
        print-color-adjust: exact !important;
    }
    
    tr.gray td {
        background: #fff !important;
        -webkit-print-color-adjust: exact;
        print-color-adjust: exact;
    }
}

img.questionPrompt { 
    max-width: 580px; 
    padding: 3px; 
    border: 1px solid #ddd; 
    display: block; 
    margin: 5px 0; 
}

img.choicePrompt { 
    max-width: 130px; 
    padding: 3px; 
    border: 1px solid #ddd; 
}

.required { 
    color: red; 
    font-style: italic; 
}

.relevance {
    font-size: 12px;
    font-style: italic;
    color: green;
}

.hint {
    color: blue;
    font-size: 13px;
}
"""

def extract_language_name(header_str):
    """Extract language identifier from column header string like 'label::English (en)' -> 'English (en)'."""
    if '::' in header_str:
        return header_str.split('::', 1)[1].strip()
    return 'Default'

def clean_language_suffix(lang_name):
    """Clean language name for filename suffix like 'English (en)' -> 'English'."""
    if not lang_name or lang_name == 'Default':
        return ''
    match = re.match(r'^([^\(\)]+)', lang_name)
    if match:
        return match.group(1).strip()
    return lang_name.strip()

class XLSFormParser:
    def __init__(self, excel_path):
        self.excel_path = excel_path
        self.wb = openpyxl.load_workbook(excel_path, data_only=True)
        self.sheets = {name.lower().strip(): name for name in self.wb.sheetnames}
        
        self.settings = self._parse_settings()
        self.languages = self._detect_languages()
        self.choices = self._parse_choices()
        self.survey_rows = self._parse_survey()

    def _get_sheet(self, target_name):
        actual_name = self.sheets.get(target_name.lower())
        if actual_name:
            return self.wb[actual_name]
        return None

    def _parse_settings(self):
        settings = {}
        sheet = self._get_sheet('settings')
        if not sheet or sheet.max_row < 2:
            return settings
        
        headers = [str(cell.value or '').strip() for cell in sheet[1]]
        values = [str(cell.value or '').strip() if cell.value is not None else '' for cell in sheet[2]]
        
        for h, v in zip(headers, values):
            if h:
                settings[h] = v
        return settings

    def _detect_languages(self):
        languages = []
        for sheet_name in ['survey', 'choices']:
            sheet = self._get_sheet(sheet_name)
            if not sheet or sheet.max_row < 1:
                continue
            headers = [str(cell.value or '').strip() for cell in sheet[1]]
            for h in headers:
                if h.startswith('label::'):
                    lang = extract_language_name(h)
                    if lang and lang not in languages:
                        languages.append(lang)
        if not languages:
            languages.append('Default')
        return languages

    def _parse_choices(self):
        choices = {}
        sheet = self._get_sheet('choices')
        if not sheet or sheet.max_row < 2:
            return choices

        headers = [str(cell.value or '').strip() for cell in sheet[1]]
        header_map = {h: i for i, h in enumerate(headers) if h}

        list_name_col = header_map.get('list_name') or header_map.get('list name')
        name_col = header_map.get('name') or header_map.get('value')

        if list_name_col is None or name_col is None:
            return choices

        for row_idx in range(2, sheet.max_row + 1):
            row_cells = sheet[row_idx]
            list_name_val = str(row_cells[list_name_col].value or '').strip()
            name_val = str(row_cells[name_col].value or '').strip()

            if not list_name_val or not name_val:
                continue

            if list_name_val not in choices:
                choices[list_name_val] = []

            labels = {}
            for lang in self.languages:
                label_header = f'label::{lang}' if lang != 'Default' else 'label'
                col_idx = header_map.get(label_header)
                if col_idx is None and lang != 'Default':
                    # Fallback matching e.g. label::English
                    for h, idx in header_map.items():
                        if h.startswith('label::') and lang.lower() in h.lower():
                            col_idx = idx
                            break
                if col_idx is None and 'label' in header_map:
                    col_idx = header_map['label']

                if col_idx is not None and col_idx < len(row_cells):
                    val = row_cells[col_idx].value
                    labels[lang] = str(val).strip() if val is not None else name_val
                else:
                    labels[lang] = name_val

            choices[list_name_val].append({
                'name': name_val,
                'labels': labels
            })

        return choices

    def _parse_survey(self):
        survey_rows = []
        sheet = self._get_sheet('survey')
        if not sheet or sheet.max_row < 2:
            return survey_rows

        headers = [str(cell.value or '').strip() for cell in sheet[1]]
        header_map = {h: i for i, h in enumerate(headers) if h}

        type_col = header_map.get('type')
        name_col = header_map.get('name')
        if type_col is None:
            return survey_rows

        for row_idx in range(2, sheet.max_row + 1):
            row_cells = sheet[row_idx]
            type_val = str(row_cells[type_col].value or '').strip()
            if not type_val:
                continue

            name_val = ''
            if name_col is not None and name_col < len(row_cells):
                name_val = str(row_cells[name_col].value or '').strip()

            # Required field checking across platform variants
            req_col = header_map.get('required')
            is_required = False
            if req_col is not None and req_col < len(row_cells):
                req_val = str(row_cells[req_col].value or '').strip().lower()
                is_required = req_val in ['yes', '1', 'true', 'ok', 'required']

            # Extract labels, hints, and constraint messages per language
            labels = {}
            hints = {}
            constraint_msgs = {}

            for lang in self.languages:
                # Label
                col_idx = header_map.get(f'label::{lang}') if lang != 'Default' else header_map.get('label')
                if col_idx is None and lang != 'Default':
                    for h, idx in header_map.items():
                        if h.startswith('label::') and lang.lower() in h.lower():
                            col_idx = idx
                            break
                if col_idx is None:
                    col_idx = header_map.get('label')
                labels[lang] = str(row_cells[col_idx].value or '').strip() if (col_idx is not None and col_idx < len(row_cells) and row_cells[col_idx].value is not None) else ''

                # Hint
                col_idx = header_map.get(f'hint::{lang}') if lang != 'Default' else header_map.get('hint')
                if col_idx is None and lang != 'Default':
                    for h, idx in header_map.items():
                        if h.startswith('hint::') and lang.lower() in h.lower():
                            col_idx = idx
                            break
                if col_idx is None:
                    col_idx = header_map.get('hint')
                hints[lang] = str(row_cells[col_idx].value or '').strip() if (col_idx is not None and col_idx < len(row_cells) and row_cells[col_idx].value is not None) else ''

                # Constraint message
                col_idx = header_map.get(f'constraint_message::{lang}') if lang != 'Default' else header_map.get('constraint_message')
                if col_idx is None and lang != 'Default':
                    for h, idx in header_map.items():
                        if h.startswith('constraint_message::') and lang.lower() in h.lower():
                            col_idx = idx
                            break
                if col_idx is None:
                    col_idx = header_map.get('constraint_message')
                constraint_msgs[lang] = str(row_cells[col_idx].value or '').strip() if (col_idx is not None and col_idx < len(row_cells) and row_cells[col_idx].value is not None) else ''

            # Logic attributes
            rel_col = header_map.get('relevant') or header_map.get('relevance')
            relevance = str(row_cells[rel_col].value or '').strip() if (rel_col is not None and rel_col < len(row_cells) and row_cells[rel_col].value is not None) else ''

            con_col = header_map.get('constraint')
            constraint = str(row_cells[con_col].value or '').strip() if (con_col is not None and con_col < len(row_cells) and row_cells[con_col].value is not None) else ''

            calc_col = header_map.get('calculation') or header_map.get('calculate')
            calculation = str(row_cells[calc_col].value or '').strip() if (calc_col is not None and calc_col < len(row_cells) and row_cells[calc_col].value is not None) else ''

            survey_rows.append({
                'type': type_val,
                'name': name_val,
                'required': is_required,
                'labels': labels,
                'hints': hints,
                'constraint_msgs': constraint_msgs,
                'relevance': relevance,
                'constraint': constraint,
                'calculation': calculation,
            })

        return survey_rows

def build_codebook_html(parser, language='Default'):
    """Build responsive HTML codebook for a specific language."""
    soup = BeautifulSoup('<!DOCTYPE html><html><head></head><body></body></html>', 'html.parser')
    
    head = soup.head
    head.append(soup.new_tag('meta', attrs={'charset': 'utf-8'}))
    head.append(soup.new_tag('meta', attrs={'name': 'viewport', 'content': 'width=device-width, initial-scale=1.0'}))
    
    title_text = parser.settings.get('form_title') or parser.settings.get('title') or os.path.splitext(os.path.basename(parser.excel_path))[0]
    title_tag = soup.new_tag('title')
    title_tag.string = title_text
    head.append(title_tag)
    
    style_tag = soup.new_tag('style')
    style_tag.string = RESPONSIVE_STYLES
    head.append(style_tag)

    body = soup.body
    container = soup.new_tag('div', attrs={'class': 'codebook-container'})
    body.append(container)

    h4 = soup.new_tag('h4')
    h4.string = title_text
    container.append(h4)

    table_wrapper = soup.new_tag('div', attrs={'class': 'table-responsive'})
    container.append(table_wrapper)

    table = soup.new_tag('table', attrs={'class': 'table table-bordered table-condensed'})
    table_wrapper.append(table)

    # Table Header
    thead = soup.new_tag('thead')
    table.append(thead)
    tr_head = soup.new_tag('tr')
    thead.append(tr_head)

    cols = [('20%', 'Field'), ('60%', 'Question'), ('20%', 'Answer')]
    for width, label in cols:
        th = soup.new_tag('th', attrs={'class': 'headerCell', 'style': f'width: {width};'})
        h6 = soup.new_tag('h6')
        h6.string = label
        th.append(h6)
        tr_head.append(th)

    tbody = soup.new_tag('tbody')
    table.append(tbody)

    for item in parser.survey_rows:
        raw_type = item['type'].strip()
        parts = raw_type.split(None, 1)
        base_type = parts[0].lower()
        list_name = parts[1].strip() if len(parts) > 1 else ''

        # Normalized section header check for ODK, SurveyCTO, KoboToolbox
        if base_type in ['begin_group', 'begin'] and raw_type.lower().startswith(('begin group', 'begin_group', 'begin repeat', 'begin_repeat')):
            label = item['labels'].get(language) or item['labels'].get('Default') or item['name']
            if not label and 'group' in raw_type.lower():
                label = item['name']
            
            tr = soup.new_tag('tr', attrs={'class': 'entryRow'})
            td = soup.new_tag('td', attrs={'colspan': '3', 'class': 'section-header-dark', 'style': 'padding-left: 5px;'})
            span = soup.new_tag('span')
            span.string = label
            td.append(span)
            tr.append(td)
            tbody.append(tr)
            continue

        if base_type in ['end_group', 'end_repeat', 'end']:
            continue

        # Standard Field Row
        tr = soup.new_tag('tr', attrs={'class': 'gray entryRow' if item['name'] else 'entryRow'})
        
        # Column 1: Field Name
        td_field = soup.new_tag('td', attrs={'class': 'fieldCell', 'style': 'padding-left: 5px;'})
        span_field = soup.new_tag('span')
        span_field.string = item['name'] if item['name'] else ''
        if item['required']:
            span_req = soup.new_tag('span', attrs={'class': 'required'})
            span_req.string = ' (required)'
            span_field.append(span_req)
        td_field.append(span_field)
        tr.append(td_field)

        # Column 2: Question Prompt & Details
        td_q = soup.new_tag('td', attrs={'class': 'questionCell'})
        q_label = item['labels'].get(language) or item['labels'].get('Default') or ''
        if q_label:
            td_q.append(q_label)

        q_hint = item['hints'].get(language) or item['hints'].get('Default') or ''
        if q_hint:
            div_hint = soup.new_tag('div', attrs={'class': 'hint'})
            div_hint.string = q_hint
            td_q.append(div_hint)

        if item['relevance']:
            div_rel = soup.new_tag('div', attrs={'class': 'relevance'})
            div_rel.string = f"Relevance: {item['relevance']}"
            td_q.append(div_rel)

        tr.append(td_q)

        # Column 3: Answer / Choice Table
        td_a = soup.new_tag('td', attrs={'class': 'answerCell'})

        if base_type in ['select_one', 'select_multiple', 'select_one_or_other', 'select_multiple_or_other'] and list_name:
            choice_list = parser.choices.get(list_name, [])
            if choice_list:
                c_table = soup.new_tag('table', attrs={'class': 'table borderless'})
                c_tbody = soup.new_tag('tbody')
                c_table.append(c_tbody)

                for choice in choice_list:
                    c_tr = soup.new_tag('tr')
                    c_td_note = soup.new_tag('td', attrs={'class': 'response-note-cell'})
                    c_td_val = soup.new_tag('td', attrs={'style': 'text-align: center; padding-left: 3px; padding-right: 3px;'})
                    c_td_val.string = choice['name']
                    
                    c_td_lbl = soup.new_tag('td', attrs={'style': 'width: 100%; padding-left: 3px; padding-right: 3px; border-left: 1px solid #999;'})
                    c_span = soup.new_tag('span')
                    c_span.string = choice['labels'].get(language) or choice['labels'].get('Default') or choice['name']
                    c_td_lbl.append(c_span)

                    c_tr.append(c_td_note)
                    c_tr.append(c_td_val)
                    c_tr.append(c_td_lbl)
                    c_tbody.append(c_tr)

                td_a.append(c_table)

        tr.append(td_a)
        tbody.append(tr)

    return str(soup)

async def generate_pdf(html_path, pdf_path):
    """Render HTML codebook to landscape A4 PDF via Playwright Chromium."""
    async with async_playwright() as p:
        browser = await p.chromium.launch()
        page = await browser.new_page()
        await page.goto(f'file:///{os.path.abspath(html_path)}', wait_until='networkidle')
        await page.wait_for_timeout(1000)
        await page.pdf(
            path=pdf_path,
            format='A4',
            landscape=True,
            margin={'top': '10mm', 'right': '10mm', 'bottom': '10mm', 'left': '10mm'},
            print_background=True
        )
        await browser.close()

def process_xlsform(excel_path, target_lang=None, generate_pdf_flag=True):
    """Process a single XLSForm spreadsheet and generate HTML/PDF codebooks."""
    print(f"Processing XLSForm: {os.path.basename(excel_path)}...")
    parser = XLSFormParser(excel_path)

    langs_to_process = parser.languages
    if target_lang:
        matching = [l for l in parser.languages if target_lang.lower() in l.lower()]
        if matching:
            langs_to_process = matching
        else:
            langs_to_process = [target_lang]

    base_name = os.path.splitext(excel_path)[0]

    for lang in langs_to_process:
        suffix = clean_language_suffix(lang)
        out_html_name = f"{base_name}_{suffix}.html" if suffix else f"{base_name}.html"
        out_pdf_name = f"{base_name}_{suffix}.pdf" if suffix else f"{base_name}.pdf"

        print(f"  Building HTML codebook for language: {lang} -> {os.path.basename(out_html_name)}")
        html_content = build_codebook_html(parser, language=lang)

        with open(out_html_name, 'w', encoding='utf-8') as f:
            f.write(html_content)
        print(f"    Saved HTML: {out_html_name}")

        if generate_pdf_flag:
            print(f"    Generating PDF: {out_pdf_name}...")
            asyncio.run(generate_pdf(out_html_name, out_pdf_name))
            print(f"    Saved PDF: {out_pdf_name}")

def main():
    parser = argparse.ArgumentParser(description="Direct XLSForm to Printable Codebook & PDF Generator")
    parser.add_argument('--input', '-i', help="Path to input XLSForm (.xlsx) file")
    parser.add_argument('--dir', '-d', help="Directory containing XLSForm files")
    parser.add_argument('--lang', '-l', help="Target language (e.g. English, Hindi). If omitted, processes all languages.")
    parser.add_argument('--no-pdf', action='store_true', help="Skip PDF generation and output HTML only.")

    args = parser.parse_args()

    if not args.input and not args.dir:
        # Default: process all .xlsx files in current directory
        cwd = os.getcwd()
        xlsx_files = [os.path.join(cwd, f) for f in os.listdir(cwd) if f.endswith('.xlsx') and not f.startswith('~$')]
        if not xlsx_files:
            print("No .xlsx files found in current directory.")
            sys.exit(1)
        for f in xlsx_files:
            process_xlsform(f, target_lang=args.lang, generate_pdf_flag=not args.no_pdf)
    elif args.input:
        process_xlsform(args.input, target_lang=args.lang, generate_pdf_flag=not args.no_pdf)
    elif args.dir:
        xlsx_files = [os.path.join(args.dir, f) for f in os.listdir(args.dir) if f.endswith('.xlsx') and not f.startswith('~$')]
        for f in xlsx_files:
            process_xlsform(f, target_lang=args.lang, generate_pdf_flag=not args.no_pdf)

if __name__ == '__main__':
    main()
